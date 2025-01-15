from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
import numpy as np
import time
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
import os
from datetime import datetime
import logging

# 配置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# 模型下载
model = SentenceTransformer('/Users/lpd/Documents/acge_text_embedding/modelscope/hub/yangjhchs/acge_text_embedding')

app = Flask(__name__)
CORS(app)  # 启用CORS
app.config['UPLOAD_FOLDER'] = 'uploads'  # 设置上传文件夹

def seconds_to_hms(seconds):
    seconds = int(seconds)
    # 转换小时
    hours = seconds // 3600
    # 计算剩余秒数
    seconds %= 3600
    # 转换分钟
    minutes = seconds // 60
    # 计算剩余秒数
    seconds %= 60
    # 返回格式化的字符串
    return f"总共用时：{hours:02d}:{minutes:02d}:{seconds:02d}"

def ragas(reference_answer, generated_answer):
    # 计算两个答案的嵌入向量
    embedding_1 = np.array(model.encode(reference_answer))
    embedding_2 = np.array(model.encode(generated_answer))

    # 计算余弦相似度
    norms_1 = np.linalg.norm(embedding_1, keepdims=True)
    norms_2 = np.linalg.norm(embedding_2, keepdims=True)
    embedding_1_normalized = embedding_1 / norms_1
    embedding_2_normalized = embedding_2 / norms_2
    similarity = embedding_1_normalized @ embedding_2_normalized.T
    score = similarity.flatten()
    similarity_score = score.tolist()[0]

    return similarity_score

def get_column_data(column_name):
    # 找到列标题对应的列索引
    column_index = None
    for col_idx, cell in enumerate(ws[1], start=1):  # 假设第一行是标题行
        if cell.value == column_name:
            column_index = col_idx
            break

    if column_index is None:
        print(f"列名 '{column_name}' 不存在于Excel文件的标题行中。")
        exit()

    # 读取整列的内容
    column_data = [ws.cell(row=row, column=column_index).value for row in range(2, ws.max_row + 1)]  # 假设从第二行开始是数据
    return column_data

def get_column_index(column_name):
    # 找到列标题对应的列索引
    column_index = None
    for col_idx, cell in enumerate(ws[1], start=1):  # 假设第一行是标题行
        if cell.value == column_name:
            column_index = col_idx
            break

    if column_index is None:
        print(f"列名 '{column_name}' 不存在于Excel文件的标题行中。")
        exit()
    return column_index

def assessz(file_path, ref, ans, data):
    # 替换为你的Excel文件路径
    similarity_column_index = get_column_index(data)
    data_to_write = []
    reference = get_column_data(ref)
    answer = get_column_data(ans)
    for index, (item1, item2) in enumerate(zip(reference, answer), start=1):
        if item1 is None:
            break
        print(f"第{index}行相似度计算")
        similarity = ragas(item1, item2)
        data_to_write.append(similarity)

    # 写入数据，从第二行开始（假设第一行是标题行）

    for row_idx, value in enumerate(data_to_write, start=2):  # 从第二行开始，行索引从2开始
        ws.cell(row=row_idx, column=similarity_column_index, value=value)

    # 保存工作簿
    wb.save(file_path)

def assess(file_path, ref, ans, data):
    # 替换为你的Excel文件路径
    question_idxs = 0
    similarity_column_index = get_column_index(data)
    data_to_write = []
    reference = get_column_data(ref)
    answer = get_column_data(ans)
    row_count = 0
    for index, (item1, item2) in enumerate(zip(reference, answer), start=1):
        if item1 is None:
            break
        print(f"第{index}行相似度计算")
        try:
            similarity = ragas(item1, item2)
            data_to_write.append(similarity)
            row_count += 1
            if row_count % batch_size == 0:  # 每50行写入一次并保存
                for row_idx, value in enumerate(data_to_write, start=index + 2 - batch_size):
                    ws.cell(row=row_idx, column=similarity_column_index, value=value)
                wb.save(file_path)
                data_to_write = []  # 重置列表，准备下一批写入
                question_idxs = index
        except Exception as e:
            print(f"处理第{index}行时发生错误: {e}")

            # 处理剩余数据（如果有）
    if data_to_write:
        for row_idx, value in enumerate(data_to_write, start=question_idxs + 2):
            ws.cell(row=row_idx, column=similarity_column_index, value=value)
        wb.save(file_path)

@app.route('/api/calculate-ass', methods=['POST'])
def process_excel():
    logger.info("收到文件上传请求")
    if 'file' not in request.files:
        logger.error("请求中没有文件")
        return jsonify({'error': '没有上传文件'}), 400
        
    file = request.files['file']
    logger.info(f"上传的文件名: {file.filename}")
    
    if not file.filename.endswith('.xlsx'):
        logger.error("文件格式不正确")
        return jsonify({'error': '只支持.xlsx格式的文件'}), 400

    try:
        # 读取Excel文件
        logger.info("开始读取Excel文件")
        wb = load_workbook(file)
        ws = wb.active
        
        # 检查工作表是否为空
        if ws.max_row < 2:  # 如果只有表头或完全为空
            logger.error("Excel文件为空或只有表头")
            return jsonify({'error': 'Excel文件为空或只有表头'}), 400
            
        # 检查工作表的列数
        if ws.max_column < 2:
            logger.error("Excel文件格式不正确：至少需要2列数据")
            return jsonify({'error': 'Excel文件格式不正确：至少需要2列数据'}), 400
        
        # 创建结果工作簿
        logger.info("创建结果工作簿")
        result_wb = Workbook()
        result_ws = result_wb.active
        
        # 设置表头
        headers = ['标准答案', '预测文本', 'ASS分数']
        for col, header in enumerate(headers, 1):
            result_ws.cell(row=1, column=col, value=header)
            
        # 处理每一行数据
        row_count = 0
        logger.info("开始处理数据")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                # 检查行是否有足够的列
                if len(row) < 2:
                    logger.warning(f"第{row_idx}行数据列数不足，跳过")
                    continue
                    
                # 检查单元格是否为空
                if not row[0].value or not row[1].value:
                    logger.warning(f"第{row_idx}行数据存在空值，跳过")
                    continue
                    
                text1 = str(row[0].value).strip()
                text2 = str(row[1].value).strip()
                
                # 检查文本是否为空字符串
                if not text1 or not text2:
                    logger.warning(f"第{row_idx}行数据存在空字符串，跳过")
                    continue
                
                # 计算相似度
                similarity = ragas(text1, text2)
                
                # 写入结果
                result_ws.cell(row=row_idx, column=1, value=text1)
                result_ws.cell(row=row_idx, column=2, value=text2)
                result_ws.cell(row=row_idx, column=3, value=similarity)
                
                row_count += 1
                
            except Exception as e:
                logger.error(f"处理第{row_idx}行时出错: {str(e)}")
                continue
            
        # 如果没有处理任何数据
        if row_count == 0:
            logger.error("文件中没有有效数据")
            return jsonify({'error': '文件中没有有效数据'}), 400
            
        # 保存结果文件
        logger.info("保存结果文件")
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        result_filename = f'ASS计算结果_{timestamp}.xlsx'
        result_filepath = os.path.join(app.config['UPLOAD_FOLDER'], result_filename)
        
        # 确保uploads目录存在
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # 保存文件
        result_wb.save(result_filepath)
        
        # 返回文件下载链接
        file_url = f'/uploads/{result_filename}'
        logger.info(f"处理完成，生成文件: {file_url}")
        response = jsonify({
            'status': 'success',
            'message': '处理完成',
            'resultFile': file_url
        })
        logger.info(f"返回响应: {response.get_data(as_text=True)}")
        return response
        
    except Exception as e:
        logger.error(f"处理文件时出错: {str(e)}", exc_info=True)
        return jsonify({'error': f'处理文件时出错: {str(e)}'}), 500

# 添加静态文件路由
@app.route('/uploads/<path:filename>')
def download_file(filename):
    try:
        logger.info(f"请求下载文件: {filename}")
        return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)
    except Exception as e:
        logger.error(f"下载文件时出错: {str(e)}", exc_info=True)
        return jsonify({'error': f'下载文件时出错: {str(e)}'}), 500

if __name__ == '__main__':
    # 记录开始时间
    start_time = time.time()
    logger.info("正在启动服务...")
    app.run(host='0.0.0.0', port=5001)
    
    # 记录结束时间并计算运行时间
    end_time = time.time()
    elapsed_time = end_time - start_time
    elapsed = seconds_to_hms(elapsed_time)
    logger.info(elapsed)
